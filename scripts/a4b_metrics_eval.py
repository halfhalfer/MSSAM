import os
from pathlib import Path

import cv2
import numpy as np
import pandas as pd
import torch
import torch.nn.functional as F
from openpyxl.utils import get_column_letter
from skimage import measure
from skimage.morphology import skeletonize
from torch.utils.data import DataLoader, Dataset
from tqdm import tqdm


class SegmentationMetricsDataset(Dataset):
    def __init__(self, pred_path, label_path, ignore_file=None, device="cpu"):
        self.pred_path = pred_path
        self.label_path = label_path
        self.device = device
        self.ignore_names = self._read_ignore_names(ignore_file)
        self.valid_pairs = self._match_pairs()

    @staticmethod
    def _read_ignore_names(ignore_file):
        ignore_names = set()
        if ignore_file and os.path.isfile(ignore_file):
            with open(ignore_file, "r", encoding="utf-8") as handle:
                for line in handle:
                    name = line.strip()
                    if name:
                        ignore_names.add(name)
        return ignore_names

    def _match_pairs(self):
        pred_files = sorted(os.listdir(self.pred_path))
        label_by_stem = {Path(name).stem: name for name in sorted(os.listdir(self.label_path))}
        valid_pairs = []
        for pred_name in pred_files:
            if pred_name in self.ignore_names:
                continue
            label_name = label_by_stem.get(Path(pred_name).stem)
            if label_name is not None:
                valid_pairs.append((pred_name, label_name))
        if not valid_pairs:
            raise ValueError("没有匹配的预测-标签文件对")
        return valid_pairs

    def __len__(self):
        return len(self.valid_pairs)

    def __getitem__(self, idx):
        pred_file, label_file = self.valid_pairs[idx]
        pred = cv2.imread(os.path.join(self.pred_path, pred_file), cv2.IMREAD_GRAYSCALE)
        label = cv2.imread(os.path.join(self.label_path, label_file), cv2.IMREAD_GRAYSCALE)
        if pred is None:
            raise FileNotFoundError(f"读取预测图失败: {os.path.join(self.pred_path, pred_file)}")
        if label is None:
            raise FileNotFoundError(f"读取标签图失败: {os.path.join(self.label_path, label_file)}")

        if pred.shape != label.shape:
            pred = cv2.resize(pred, (label.shape[1], label.shape[0]), interpolation=cv2.INTER_NEAREST)

        if np.issubdtype(pred.dtype, np.floating):
            pred = (pred > 0.5).astype(np.uint8)
        else:
            pred = (pred > 127).astype(np.uint8)
        label = np.where((label == 1) | (label == 255), 1, 0).astype(np.uint8)

        return {
            "pred": torch.from_numpy(pred).to(self.device),
            "label": torch.from_numpy(label).to(self.device),
            "filename": pred_file,
        }


def calculate_single_seg_metrics(pred, label, epsilon=1e-8):
    pred_flat = pred.flatten()
    label_flat = label.flatten()

    tp = torch.sum((pred_flat == 1) & (label_flat == 1)).item()
    tn = torch.sum((pred_flat == 0) & (label_flat == 0)).item()
    fp = torch.sum((pred_flat == 1) & (label_flat == 0)).item()
    fn = torch.sum((pred_flat == 0) & (label_flat == 1)).item()

    recall_0 = tn / (tn + fp + epsilon) if (tn + fp) > 0 else 0
    precision_0 = tn / (tn + fn + epsilon) if (tn + fn) > 0 else 0
    f1_0 = 2 * (precision_0 * recall_0) / (precision_0 + recall_0 + epsilon) if (precision_0 + recall_0) > 0 else 0

    recall_1 = tp / (tp + fn + epsilon) if (tp + fn) > 0 else 0
    precision_1 = tp / (tp + fp + epsilon) if (tp + fp) > 0 else 0
    f1_1 = 2 * (precision_1 * recall_1) / (precision_1 + recall_1 + epsilon) if (precision_1 + recall_1) > 0 else 0

    mean_recall = (recall_0 + recall_1) / 2
    mean_precision = (precision_0 + precision_1) / 2
    mean_f1 = (f1_0 + f1_1) / 2

    iou_0 = tn / (tn + fp + fn + epsilon) if (tn + fp + fn) > 0 else 0
    iou_1 = tp / (tp + fp + fn + epsilon) if (tp + fp + fn) > 0 else 0
    miou = (iou_0 + iou_1) / 2

    total = tp + tn + fp + fn
    po = (tp + tn) / (total + epsilon)
    pe_num = ((tp + fp) * (tp + fn)) + ((fn + tn) * (fp + tn))
    pe = pe_num / (total**2 + epsilon)
    kappa = (po - pe) / (1 - pe + epsilon) if (1 - pe) > 0 else 0
    accuracy = (tp + tn) / (total + epsilon)

    return {
        "TP": tp,
        "TN": tn,
        "FP": fp,
        "FN": fn,
        "Recall": mean_recall,
        "Precision": mean_precision,
        "F1": mean_f1,
        "Kappa": kappa,
        "mIoU": miou,
        "Accuracy": accuracy,
        "Recall_0": recall_0,
        "Precision_0": precision_0,
        "F1_0": f1_0,
        "IoU_0": iou_0,
        "Recall_1": recall_1,
        "Precision_1": precision_1,
        "F1_1": f1_1,
        "IoU_1": iou_1,
    }


def generate_segmentation_report(pred_path, label_path, ignore_file=None, batch_size=4, device="cpu"):
    dataset = SegmentationMetricsDataset(pred_path, label_path, ignore_file=ignore_file, device=device)
    dataloader = DataLoader(dataset, batch_size=batch_size, shuffle=False)

    results = []
    with tqdm(total=len(dataset), desc="语义指标计算中", unit="img") as pbar:
        for batch in dataloader:
            curr_batch_size = batch["pred"].shape[0]
            for i in range(curr_batch_size):
                metrics = calculate_single_seg_metrics(batch["pred"][i], batch["label"][i])
                metrics["Filename"] = batch["filename"][i]
                results.append(metrics)
                pbar.update(1)

    total_tp = sum(m["TP"] for m in results)
    total_tn = sum(m["TN"] for m in results)
    total_fp = sum(m["FP"] for m in results)
    total_fn = sum(m["FN"] for m in results)
    epsilon = 1e-8
    tp, tn, fp, fn = total_tp, total_tn, total_fp, total_fn
    total = tp + tn + fp + fn

    overall_recall_0 = tn / (tn + fp + epsilon) if (tn + fp) > 0 else 0
    overall_precision_0 = tn / (tn + fn + epsilon) if (tn + fn) > 0 else 0
    overall_f1_0 = 2 * (overall_precision_0 * overall_recall_0) / (overall_precision_0 + overall_recall_0 + epsilon) if (overall_precision_0 + overall_recall_0) > 0 else 0
    overall_recall_1 = tp / (tp + fn + epsilon) if (tp + fn) > 0 else 0
    overall_precision_1 = tp / (tp + fp + epsilon) if (tp + fp) > 0 else 0
    overall_f1_1 = 2 * (overall_precision_1 * overall_recall_1) / (overall_precision_1 + overall_recall_1 + epsilon) if (overall_precision_1 + overall_recall_1) > 0 else 0

    iou_0 = tn / (tn + fp + fn + epsilon) if (tn + fp + fn) > 0 else 0
    iou_1 = tp / (tp + fp + fn + epsilon) if (tp + fp + fn) > 0 else 0
    po = (tp + tn) / (total + epsilon)
    pe_num = ((tp + fp) * (tp + fn)) + ((fn + tn) * (fp + tn))
    pe = pe_num / (total**2 + epsilon)

    results.append(
        {
            "Filename": "OVERALL",
            "Recall": (overall_recall_0 + overall_recall_1) / 2,
            "Precision": (overall_precision_0 + overall_precision_1) / 2,
            "F1": (overall_f1_0 + overall_f1_1) / 2,
            "Kappa": (po - pe) / (1 - pe + epsilon) if (1 - pe) > 0 else 0,
            "mIoU": (iou_0 + iou_1) / 2,
            "Accuracy": (tp + tn) / (total + epsilon),
            "TP": tp,
            "TN": tn,
            "FP": fp,
            "FN": fn,
            "Recall_0": overall_recall_0,
            "Precision_0": overall_precision_0,
            "F1_0": overall_f1_0,
            "IoU_0": iou_0,
            "Recall_1": overall_recall_1,
            "Precision_1": overall_precision_1,
            "F1_1": overall_f1_1,
            "IoU_1": iou_1,
        }
    )
    return pd.DataFrame(results)


def calculate_edge_details(pred, label, buffer_size=3):
    p = pred.float().view(1, 1, pred.shape[-2], pred.shape[-1])
    l = label.float().view(1, 1, label.shape[-2], label.shape[-1])

    def get_buffer(mask, b):
        if b <= 0:
            return mask
        return F.max_pool2d(mask, kernel_size=2 * b + 1, stride=1, padding=b)

    com_den = l.sum().item()
    p_buffer = get_buffer(p, buffer_size)
    com_num = (l * p_buffer).sum().item()
    com = com_num / (com_den + 1e-8) if com_den > 0 else 0

    corr_den = p.sum().item()
    l_buffer = get_buffer(l, buffer_size)
    corr_num = (p * l_buffer).sum().item()
    corr = corr_num / (corr_den + 1e-8) if corr_den > 0 else 0

    f_edge = 2 * (com * corr) / (com + corr + 1e-8) if (com + corr) > 0 else 0
    return {
        "L_truth_len": com_den,
        "L_pred_match": com_num,
        "L_pred_len": corr_den,
        "L_truth_match": corr_num,
        "Completeness": com,
        "Correctness": corr,
        "F_edge": f_edge,
    }


def calculate_edge_details_v2(pred, label, buffer_size=3, a4b_fix=False):
    p_raw = pred.float().view(1, 1, pred.shape[-2], pred.shape[-1])
    l = label.float().view(1, 1, label.shape[-2], label.shape[-1])
    device = pred.device

    p_np = (p_raw.squeeze() > 0.5).cpu().numpy().astype(np.uint8)
    if p_np.sum() > 0:
        p_skel = torch.from_numpy(skeletonize(p_np).astype(np.float32)).to(device).view(1, 1, *p_np.shape)
    else:
        p_skel = p_raw.clone()

    def get_buffer(mask, b):
        if b <= 0:
            return mask
        return F.max_pool2d(mask, kernel_size=2 * b + 1, stride=1, padding=b)

    com_den = l.sum().item()
    p_buffer = get_buffer(p_skel, buffer_size)
    com_num = (l * p_buffer).sum().item()
    com = com_num / (com_den + 1e-8) if com_den > 0 else 0

    corr_den = p_raw.sum().item()
    l_buffer = get_buffer(l, 0 if a4b_fix else buffer_size)
    corr_num = (p_raw * l_buffer).sum().item()
    corr = corr_num / (corr_den + 1e-8) if corr_den > 0 else 0

    f_edge = 2 * (com * corr) / (com + corr + 1e-8) if (com + corr) > 0 else 0
    return {
        "L_truth_len": com_den,
        "L_pred_match": com_num,
        "L_pred_len": corr_den,
        "L_truth_match": corr_num,
        "Completeness": com,
        "Correctness": corr,
        "F_edge": f_edge,
    }


def generate_edge_report(pred_path, label_path, ignore_file=None, batch_size=4, device="cpu", buffer_size=1, a4b_fix=False, edge_metric_v=2):
    dataset = SegmentationMetricsDataset(pred_path, label_path, ignore_file=ignore_file, device=device)
    dataloader = DataLoader(dataset, batch_size=batch_size, shuffle=False)

    results = []
    with tqdm(total=len(dataset), desc="边缘指标计算中", unit="img") as pbar:
        for batch in dataloader:
            curr_batch_size = batch["pred"].shape[0]
            for i in range(curr_batch_size):
                if edge_metric_v == 1:
                    metrics = calculate_edge_details(batch["pred"][i], batch["label"][i], buffer_size=buffer_size)
                else:
                    metrics = calculate_edge_details_v2(
                        batch["pred"][i],
                        batch["label"][i],
                        buffer_size=buffer_size,
                        a4b_fix=a4b_fix,
                    )
                metrics["Filename"] = batch["filename"][i]
                results.append(metrics)
                pbar.update(1)

    df = pd.DataFrame(results)
    sum_l_truth = df["L_truth_len"].sum()
    sum_l_pred_match = df["L_pred_match"].sum()
    sum_l_pred = df["L_pred_len"].sum()
    sum_l_truth_match = df["L_truth_match"].sum()
    overall_com = sum_l_pred_match / (sum_l_truth + 1e-8)
    overall_corr = sum_l_truth_match / (sum_l_pred + 1e-8)
    overall_fedge = 2 * (overall_com * overall_corr) / (overall_com + overall_corr + 1e-8)
    overall_row = {
        "Filename": "OVERALL_TOTAL",
        "L_truth_len": sum_l_truth,
        "L_pred_match": sum_l_pred_match,
        "L_pred_len": sum_l_pred,
        "L_truth_match": sum_l_truth_match,
        "Completeness": overall_com,
        "Correctness": overall_corr,
        "F_edge": overall_fedge,
    }
    df = pd.concat([df, pd.DataFrame([overall_row])], ignore_index=True)
    cols = ["Filename", "Completeness", "Correctness", "F_edge", "L_truth_len", "L_pred_match", "L_pred_len", "L_truth_match"]
    return df[cols]


def compute_object_metrics_from_instance(pred_instance, gt_instance, use_weight=False, weight_mode="gt", return_detail=False):
    pred_regions = measure.regionprops(pred_instance)
    gt_regions = measure.regionprops(gt_instance)
    gt_by_label = {region.label: region for region in gt_regions}

    oc_list, uc_list, tc_list, pred_areas, gt_areas = [], [], [], [], []
    for pred_region in pred_regions:
        pred_mask = pred_instance == pred_region.label
        overlap_gt = gt_instance[pred_mask]
        unique_gt = np.unique(overlap_gt[overlap_gt != 0])

        if len(unique_gt) == 0:
            oc, uc, area_gt = 1.0, 1.0, 0
        else:
            overlap_areas = [np.sum(overlap_gt == gid) for gid in unique_gt]
            area_overlap = np.max(overlap_areas)
            max_gt_id = unique_gt[np.argmax(overlap_areas)]
            gt_region = gt_by_label.get(max_gt_id)
            area_gt = gt_region.area if gt_region is not None else 0
            area_pred = pred_region.area
            oc = 1 - area_overlap / area_gt if area_gt > 0 else 1.0
            uc = 1 - area_overlap / area_pred if area_pred > 0 else 1.0

        tc = np.sqrt(oc * uc)
        oc_list.append(oc)
        uc_list.append(uc)
        tc_list.append(tc)
        pred_areas.append(pred_region.area)
        gt_areas.append(area_gt)

    oc_list = np.array(oc_list)
    uc_list = np.array(uc_list)
    tc_list = np.array(tc_list)
    pred_areas = np.array(pred_areas)
    gt_areas = np.array(gt_areas)

    if use_weight and len(oc_list) > 0:
        if weight_mode == "gt":
            weights = gt_areas / np.sum(gt_areas + 1e-10)
        else:
            weights = pred_areas / np.sum(pred_areas + 1e-10)
        goc = np.sum(oc_list * weights)
        guc = np.sum(uc_list * weights)
        gtc = np.sum(tc_list * weights)
    else:
        goc = np.mean(oc_list) if len(oc_list) > 0 else 0
        guc = np.mean(uc_list) if len(uc_list) > 0 else 0
        gtc = np.mean(tc_list) if len(tc_list) > 0 else 0

    if return_detail:
        return goc, guc, gtc, oc_list, uc_list, tc_list, pred_areas, gt_areas
    return goc, guc, gtc


def aggregate_object_metrics(all_oc, all_uc, all_tc, all_areas, per_image_means):
    total_area = np.sum(all_areas) if len(all_areas) > 0 else 1e-10
    return {
        "macro": {
            "GOC": np.nanmean([item["GOC"] for item in per_image_means]),
            "GUC": np.nanmean([item["GUC"] for item in per_image_means]),
            "GTC": np.nanmean([item["GTC"] for item in per_image_means]),
        },
        "area_weighted": {
            "GOC": np.sum(all_oc * all_areas) / total_area,
            "GUC": np.sum(all_uc * all_areas) / total_area,
            "GTC": np.sum(all_tc * all_areas) / total_area,
        },
        "micro": {
            "GOC": np.sum(all_oc * all_areas) / total_area,
            "GUC": np.sum(all_uc * all_areas) / total_area,
            "GTC": np.sum(all_tc * all_areas) / total_area,
        },
    }


class InstanceSegmentationEvaluator:
    def __init__(self, pred_path, label_path, iou_thresh=0.5, cloud_txt=None):
        self.pred_path = pred_path
        self.label_path = label_path
        self.iou_thresh = iou_thresh
        self.cloud_samples = self._read_cloud_txt(cloud_txt) if cloud_txt else set()
        self.valid_pairs = self._match_pairs()

    @staticmethod
    def _read_cloud_txt(cloud_txt):
        with open(cloud_txt, "r", encoding="utf-8") as handle:
            return set(handle.read().splitlines())

    def _match_pairs(self):
        pred_files = sorted(os.listdir(self.pred_path))
        label_by_stem = {Path(name).stem: name for name in sorted(os.listdir(self.label_path))}
        valid_pairs = []
        for pred_name in pred_files:
            if pred_name in self.cloud_samples:
                continue
            label_name = label_by_stem.get(Path(pred_name).stem)
            if label_name is not None:
                valid_pairs.append((pred_name, label_name))
        if not valid_pairs:
            raise ValueError("没有匹配的预测-标签文件对")
        return valid_pairs

    @staticmethod
    def _read_mask(path):
        mask = cv2.imread(path, cv2.IMREAD_UNCHANGED)
        if mask is None:
            raise FileNotFoundError(f"读取失败：{path}")
        return mask

    def evaluate_single_torch(self, gt_mask_np, pred_mask_np, device="cpu"):
        gt_mask = torch.from_numpy(gt_mask_np.astype(np.int32)).to(device)
        pred_mask = torch.from_numpy(pred_mask_np.astype(np.int32)).to(device)
        gt_ids = torch.unique(gt_mask)
        pred_ids = torch.unique(pred_mask)
        gt_ids = gt_ids[gt_ids != 0]
        pred_ids = pred_ids[pred_ids != 0]

        num_gt = len(gt_ids)
        num_pred = len(pred_ids)
        if num_gt == 0 and num_pred == 0:
            return {"TP": 0, "FP": 0, "FN": 0, "Precision": 1.0, "Recall": 1.0, "F1": 1.0}
        if num_gt == 0 and num_pred > 0:
            return {"TP": 0, "FP": num_pred, "FN": 0, "Precision": 0.0, "Recall": 0.0, "F1": 0.0}
        if num_gt > 0 and num_pred == 0:
            return {"TP": 0, "FP": 0, "FN": num_gt, "Precision": 0.0, "Recall": 0.0, "F1": 0.0}

        pred_masks = pred_mask[None, :, :] == pred_ids[:, None, None]
        gt_masks = gt_mask[None, :, :] == gt_ids[:, None, None]
        pred_flat = pred_masks.view(num_pred, -1).float()
        gt_flat = gt_masks.view(num_gt, -1).float()
        intersection = torch.matmul(pred_flat, gt_flat.T)
        pred_sum = pred_flat.sum(dim=1, keepdim=True)
        gt_sum = gt_flat.sum(dim=1, keepdim=True).T
        union = pred_sum + gt_sum - intersection
        iou_matrix = intersection / (union + 1e-6)

        tp = 0
        fp = 0
        matched_gt = set()
        for i in range(num_pred):
            best_gt = torch.argmax(iou_matrix[i]).item()
            best_iou = iou_matrix[i, best_gt].item()
            if best_iou >= self.iou_thresh and best_gt not in matched_gt:
                tp += 1
                matched_gt.add(best_gt)
            else:
                fp += 1
        fn = num_gt - len(matched_gt)
        precision = tp / (tp + fp + 1e-8)
        recall = tp / (tp + fn + 1e-8)
        f1 = 2 * precision * recall / (precision + recall + 1e-8)
        return {"TP": tp, "FP": fp, "FN": fn, "Precision": precision, "Recall": recall, "F1": f1}

    def generate_report(self, use_weight=True, weight_mode="gt"):
        results = []
        all_tp, all_fp, all_fn = 0, 0, 0
        all_oc, all_uc, all_tc, all_areas = [], [], [], []
        gt_areas_list = []
        per_image_means = []

        for pred_file, label_file in tqdm(self.valid_pairs, desc="实例指标计算中", unit="img"):
            pred_mask = self._read_mask(os.path.join(self.pred_path, pred_file))
            gt_mask = self._read_mask(os.path.join(self.label_path, label_file))
            if pred_mask.shape != gt_mask.shape:
                pred_mask = cv2.resize(pred_mask, (gt_mask.shape[1], gt_mask.shape[0]), interpolation=cv2.INTER_NEAREST)

            metrics = self.evaluate_single_torch(gt_mask, pred_mask)
            metrics["Filename"] = pred_file
            all_tp += metrics["TP"]
            all_fp += metrics["FP"]
            all_fn += metrics["FN"]

            goc, guc, gtc, oc_list, uc_list, tc_list, areas, gt_areas = compute_object_metrics_from_instance(
                pred_mask,
                gt_mask,
                use_weight=use_weight,
                return_detail=True,
            )
            metrics["GOC"] = goc
            metrics["GUC"] = guc
            metrics["GTC"] = gtc
            results.append(metrics)
            all_oc.extend(oc_list)
            all_uc.extend(uc_list)
            all_tc.extend(tc_list)
            all_areas.extend(areas)
            gt_areas_list.extend(gt_areas)
            per_image_means.append({"GOC": goc, "GUC": guc, "GTC": gtc})

        precision = all_tp / (all_tp + all_fp + 1e-8)
        recall = all_tp / (all_tp + all_fn + 1e-8)
        f1 = 2 * precision * recall / (precision + recall + 1e-8)
        area_source = np.array(all_areas if weight_mode == "gt" else gt_areas_list)
        dataset_metrics = aggregate_object_metrics(np.array(all_oc), np.array(all_uc), np.array(all_tc), area_source, per_image_means)

        for mode_name, vals in dataset_metrics.items():
            results.append(
                {
                    "Filename": f"OVERALL_{mode_name.upper()}",
                    "TP": all_tp,
                    "FP": all_fp,
                    "FN": all_fn,
                    "Precision": precision,
                    "Recall": recall,
                    "F1": f1,
                    "GOC": vals["GOC"],
                    "GUC": vals["GUC"],
                    "GTC": vals["GTC"],
                }
            )
        return pd.DataFrame(results)

    @staticmethod
    def save_report(df, save_path):
        df.to_excel(save_path, index=False)


def save_excel_report_v2(df, save_path):
    columns = [
        "Filename",
        "Accuracy",
        "Recall",
        "Precision",
        "F1",
        "Recall_0",
        "Precision_0",
        "F1_0",
        "IoU_0",
        "Recall_1",
        "Precision_1",
        "F1_1",
        "IoU_1",
        "mIoU",
        "Kappa",
        "TP",
        "TN",
        "FP",
        "FN",
    ]
    df = df.reindex(columns=columns)
    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        sheet_name = "Results"
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        for col_idx, col in enumerate(df.columns, start=1):
            if col != "Filename":
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, (int, float)) and pd.notna(cell.value):
                        cell.number_format = "0.0000"
        for i, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[get_column_letter(i)].width = max_len


def save_edge_excel_report_com_corr(df, save_path):
    columns = ["Filename", "Completeness", "Correctness", "F_edge", "L_truth_len", "L_pred_match", "L_pred_len", "L_truth_match"]
    df = df.reindex(columns=[col for col in columns if col in df.columns])
    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        sheet_name = "Edge_Evaluation"
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        for col_idx, col_name in enumerate(df.columns, start=1):
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)) and pd.notna(cell.value):
                    if col_name in ["Completeness", "Correctness", "F_edge"]:
                        cell.number_format = "0.0000"
                    elif "len" in col_name or "match" in col_name:
                        cell.number_format = "0"
        for i, col in enumerate(df.columns, start=1):
            max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 3
            worksheet.column_dimensions[get_column_letter(i)].width = max_len


def ensure_report_dirs(out_p):
    for folder in ["Seg", "Edge", "Instance"]:
        os.makedirs(os.path.join(out_p, folder), exist_ok=True)


def gen_metics(in_sem_p, in_edge_p, in_instance_p, label_sem_p, label_edge_p, label_instance_p, out_p, dilate_iter=0, file_name="default", use_weight=True, cloud_txt=None):
    seg_df = generate_segmentation_report(in_sem_p, label_sem_p, device="cpu", ignore_file=cloud_txt)
    edge_df = generate_edge_report(in_edge_p, label_edge_p, device="cpu", buffer_size=dilate_iter, ignore_file=cloud_txt, edge_metric_v=1)
    instance_df = InstanceSegmentationEvaluator(in_instance_p, label_instance_p, cloud_txt=cloud_txt).generate_report(use_weight=use_weight)

    ensure_report_dirs(out_p)
    save_excel_report_v2(seg_df, os.path.join(out_p, "Seg", f"{file_name}.xlsx"))
    save_excel_report_v2(edge_df, os.path.join(out_p, "Edge", f"{file_name}_dilate_{dilate_iter}.xlsx"))
    InstanceSegmentationEvaluator.save_report(instance_df, os.path.join(out_p, "Instance", f"{file_name}.xlsx"))


def gen_metics_v2(in_sem_p, in_edge_p, in_instance_p, label_sem_p, label_edge_p, label_instance_p, out_p, dilate_iter=0, file_name="default", use_weight=True, cloud_txt=None, edge_batch_size=4, edge_device="cpu", edge_metric_v=2, a4b_fix=False):
    seg_df = generate_segmentation_report(in_sem_p, label_sem_p, device="cpu", ignore_file=cloud_txt)
    edge_df = generate_edge_report(
        in_edge_p,
        label_edge_p,
        ignore_file=cloud_txt,
        batch_size=edge_batch_size,
        device=edge_device,
        buffer_size=dilate_iter,
        a4b_fix=a4b_fix,
        edge_metric_v=edge_metric_v,
    )
    instance_df = InstanceSegmentationEvaluator(in_instance_p, label_instance_p, cloud_txt=cloud_txt).generate_report(use_weight=use_weight)

    ensure_report_dirs(out_p)
    save_excel_report_v2(seg_df, os.path.join(out_p, "Seg", f"{file_name}.xlsx"))
    save_edge_excel_report_com_corr(edge_df, os.path.join(out_p, "Edge", f"{file_name}_buffer_{dilate_iter}.xlsx"))
    InstanceSegmentationEvaluator.save_report(instance_df, os.path.join(out_p, "Instance", f"{file_name}.xlsx"))


def gen_metics_v3(in_sem_p, in_edge_p, in_instance_p, label_sem_p, label_edge_p, label_instance_p, out_p, dilate_iter=0, file_name="default", use_weight=True, cloud_txt=None, edge_batch_size=4, edge_device="cpu", edge_metric_v=2, a4b_fix=False):
    seg_df = generate_segmentation_report(in_sem_p, label_sem_p, device="cpu", ignore_file=cloud_txt)
    edge_df = generate_edge_report(
        in_edge_p,
        label_edge_p,
        ignore_file=cloud_txt,
        batch_size=edge_batch_size,
        device=edge_device,
        buffer_size=dilate_iter,
        a4b_fix=a4b_fix,
        edge_metric_v=edge_metric_v,
    )
    instance_df = InstanceSegmentationEvaluator(in_instance_p, label_instance_p, cloud_txt=cloud_txt).generate_report(use_weight=use_weight)

    ensure_report_dirs(out_p)
    save_excel_report_v2(seg_df, os.path.join(out_p, "Seg", f"{file_name}.xlsx"))
    save_edge_excel_report_com_corr(edge_df, os.path.join(out_p, "Edge", f"{file_name}_dilate_{dilate_iter}.xlsx"))
    InstanceSegmentationEvaluator.save_report(instance_df, os.path.join(out_p, "Instance", f"{file_name}.xlsx"))

    seg_core = seg_df.iloc[-1:][["Accuracy", "F1", "mIoU"]].copy()
    edge_core = edge_df.iloc[-1:][["Completeness", "Correctness", "F_edge"]].copy()
    ins_core = instance_df.iloc[-1:][["GOC", "GUC", "GTC"]].copy()
    summary_row = pd.concat(
        [seg_core.reset_index(drop=True), edge_core.reset_index(drop=True), ins_core.reset_index(drop=True)],
        axis=1,
    )
    summary_row.insert(0, "Model_Seed", file_name)
    return summary_row


def write_smoke_summary(summary_df, out_dir, file_name):
    summary_path = os.path.join(out_dir, f"{file_name}_summary.csv")
    summary_df.to_csv(summary_path, index=False)
    txt_path = os.path.join(out_dir, f"{file_name}_summary.txt")
    with open(txt_path, "w", encoding="utf-8") as handle:
        handle.write(summary_df.to_string(index=False))
        handle.write("\n")
    return summary_path, txt_path


if __name__ == "__main__":
    project_root = Path(__file__).resolve().parent.parent
    smoke_pred_root = project_root / "smoke_runs" / "a4b_existing_ckpt_test"
    smoke_out_root = smoke_pred_root / "smokerun" / "metrics_a4b_eval"

    label_sem_path = "/mnt/disk3/har/DataSet/Cropland/Ai4Boundaries/Processed_data/mask/test/Semantic"
    label_edge_path = "/mnt/disk3/har/DataSet/Cropland/Ai4Boundaries/Processed_data/mask/test/Edge"
    label_instance_path = "/mnt/disk3/har/DataSet/Cropland/Ai4Boundaries/Processed_data/mask/test/channel_3"
    cloud_txt = "/mnt/disk3/har/DataSet/Cropland/Ai4Boundaries/Processed_data/test_cloud_image_list_v2.txt"

    smoke_out_root.mkdir(parents=True, exist_ok=True)
    summary_df = gen_metics_v3(
        in_sem_p=str(smoke_pred_root / "test_hq_mask"),
        in_edge_p=str(smoke_pred_root / "test_hq_edge"),
        in_instance_p=str(smoke_pred_root / "parcels" / "bin"),
        label_sem_p=label_sem_path,
        label_edge_p=label_edge_path,
        label_instance_p=label_instance_path,
        out_p=str(smoke_out_root),
        dilate_iter=1,
        file_name="a4b_existing_ckpt_test",
        use_weight=True,
        cloud_txt=cloud_txt,
        edge_batch_size=4,
        edge_device="cpu",
        edge_metric_v=2,
        a4b_fix=False,
    )
    csv_path, txt_path = write_smoke_summary(summary_df, str(smoke_out_root), "a4b_existing_ckpt_test")
    print(summary_df.to_string(index=False))
    print(f"Summary CSV saved to: {csv_path}")
    print(f"Summary TXT saved to: {txt_path}")
